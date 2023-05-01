from django.shortcuts import render

# views.py

# from django.http import HttpResponse
# from django.views.decorators.csrf import csrf_exempt
# from rest_framework.decorators import api_view
# from rest_framework.response import Response
# from openpyxl import load_workbook
# from reportlab.lib.pagesizes import letter
# from reportlab.pdfgen import canvas


# @csrf_exempt
# @api_view(['POST'])
# def convert_excel_to_pdf(request):
#     if request.method == 'POST':
#         # Load the Excel file from the request
#         excel_file = request.FILES.get('file')
#         wb = load_workbook(excel_file)
#         ws = wb.active
        
#         # Create a PDF file using reportlab
#         response = HttpResponse(content_type='application/pdf')
#         response['Content-Disposition'] = 'attachment; filename="output.pdf"'
#         pdf = canvas.Canvas(response, pagesize=letter)
        
#         # Write the contents of the Excel file to the PDF
#         for row in ws.rows:
#             for cell in row:
#                 pdf.drawString(cell.column_letter + str(cell.row), str(cell.value))
                
#         # Close the PDF and return it in the response
#         pdf.save()
#         return response
# import io
# import os
# from django.http import HttpResponse
# from django.shortcuts import render
# from django.template.loader import get_template
# from django.views.decorators.csrf import csrf_exempt
# from openpyxl import load_workbook
# from reportlab.lib.pagesizes import letter
# from reportlab.pdfgen import canvas

# def convert_excel_to_pdf(file):
#     # Load the Excel file into memory
#     workbook = load_workbook(file, data_only=True)
#     worksheet = workbook.active
    
#     # Create a new PDF file in memory
#     pdf_buffer = io.BytesIO()
#     pdf = canvas.Canvas(pdf_buffer, pagesize=letter)
    
#     # Iterate over the rows and columns of the Excel worksheet and draw them onto the PDF canvas
#     for row in worksheet.iter_rows():
#         for cell in row:
#             pdf.drawString(cell.column_letter, 792 - cell.row * 11, str(cell.value))
    
#     # Save the PDF file and return it
#     pdf.save()
#     pdf_buffer.seek(0)
#     return pdf_buffer

# @csrf_exempt
# def convert(request):
#     if request.method == 'POST':
#         file = request.FILES['file']
#         pdf_buffer = convert_excel_to_pdf(file)
#         response = HttpResponse(pdf_buffer, content_type='application/pdf')
#         response['Content-Disposition'] = 'attachment; filename="output.pdf"'
#         return response
#     else:
#         return render(request, 'convert.html')

# import io
# from django.http import HttpResponse
# from django.shortcuts import render
# from django.template.loader import get_template
# from django.views.decorators.csrf import csrf_exempt
# from openpyxl import load_workbook
# from reportlab.lib.pagesizes import letter
# from reportlab.pdfgen import canvas

# @csrf_exempt
# def convert_excel_to_pdf(request):
#     if request.method == 'POST':
#         file = request.FILES['file']
#         file_buffer = io.BytesIO(file.read())
#         file_buffer.seek(0)
        
#         # Load the Excel file into memory
#         workbook = load_workbook(file_buffer, data_only=True)
#         worksheet = workbook.active
#         # Convert all cell values to strings
        

        
#         # Create a new PDF file in memory
#         pdf_buffer = io.BytesIO()
#         pdf = canvas.Canvas(pdf_buffer, pagesize=letter)
        
#         # Iterate over the rows and columns of the Excel worksheet and draw them onto the PDF canvas
#         for row in worksheet.iter_rows():
#             for cell in row:
#                 pdf.drawString(cell.column_letter, 792 - cell.row * 11, str(cell.value))


#         for row in worksheet.iter_rows():
#             for cell in row:
#                 if isinstance(cell.value, (int, float)):
#                     pdf.drawString(cell.column_letter, 792 - cell.row * 11, str(cell.value))
                
#                 else:
#                     pdf.drawString(cell.column_letter, 792 - cell.row * 11, "N/A")


        
#         # Save the PDF file and return it
#         pdf.save()
#         pdf_buffer.seek(0)
        
#         response = HttpResponse(pdf_buffer, content_type='application/pdf')
#         response['Content-Disposition'] = 'attachment; filename="output.pdf"'
#         return response
#     else:
#         return render(request, 'convert.html')

import io
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

@csrf_exempt
def convert_excel_to_pdf(request):
    if request.method == 'POST':
        file = request.FILES['file']
        file_buffer = io.BytesIO(file.read())
        file_buffer.seek(0)
        
        # Load the Excel file into memory
        workbook = load_workbook(file_buffer, data_only=True)
        worksheet = workbook.active
        
        # Create a new PDF file in memory
        pdf_buffer = io.BytesIO()
        pdf = canvas.Canvas(pdf_buffer, pagesize=letter)
        
        # Set font properties
        pdf.setFont('Helvetica', 12)
        
        # Iterate over the rows and columns of the Excel worksheet and draw them onto the PDF canvas
        for row in worksheet.iter_rows():
            for cell in row:
                # Convert None values to empty strings
                if cell.value is None:
                    value = ''
                else:
                    value = str(cell.value)
                
                # Draw the cell onto the PDF canvas
                pdf.drawString(cell.column_letter, 792 - cell.row * 11, value)
        
        # Save the PDF file and return it
        pdf.save()
        pdf_buffer.seek(0)
        
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="output.pdf"'
        return response
    else:
        return render(request, 'convert.html')
