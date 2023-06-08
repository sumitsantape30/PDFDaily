from django.shortcuts import render

# Create your views here.
import os
import tempfile
from django.http import HttpResponse
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from rest_framework.decorators import api_view
from rest_framework.response import Response

@api_view(['POST'])
def excel_to_pdf(request):
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return HttpResponse("No file submitted.", status=400)

    # Save the uploaded Excel file temporarily
    with tempfile.NamedTemporaryFile(delete=False) as temp_excel:
        temp_excel.write(excel_file.read())

    # Load the Excel file using openpyxl
    wb = load_workbook(temp_excel.name)
    ws = wb.active

    # Define the output PDF file path
    temp_pdf_path = os.path.join(tempfile.gettempdir(), 'temp_pdf.pdf')

    # Convert Excel to PDF using reportlab
    pdf_canvas = canvas.Canvas(temp_pdf_path)
    for row in ws.iter_rows():
        for cell in row:
            pdf_canvas.drawString(cell.column_letter, ws.max_row - cell.row + 1, str(cell.value))
    pdf_canvas.save()

    # Serve the converted PDF file as a downloadable response
    with open(temp_pdf_path, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="converted_file.pdf"'

    # Delete temporary files
    os.remove(temp_excel.name)
    os.remove(temp_pdf_path)

    return response
